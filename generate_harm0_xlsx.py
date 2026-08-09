# -*- coding: utf-8 -*-
"""扫描所有角色文件夹，取每张 *-Harm_0.png 图片，生成包含图片的 XLSX 表格。"""
import os
import glob
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter

ROOT = r"e:\BaiduNetdiskDownload\新版本"
OUTPUT = os.path.join(ROOT, "角色形象对照表.xlsx")

# 收集所有子文件夹
folders = sorted(
    [d for d in os.listdir(ROOT)
     if os.path.isdir(os.path.join(ROOT, d)) and not d.startswith(".")]
)

rows = []          # (文件夹名, 图片路径)
missing = []       # 没有找到 Harm_0 的文件夹

for folder in folders:
    folder_path = os.path.join(ROOT, folder)
    # 特殊：veronica.png 文件夹内直接放了一张 veronica.png 图片
    inner = os.path.join(folder_path, "veronica.png")
    if os.path.isfile(inner):
        rows.append((folder, inner))
        continue
    png_dir = os.path.join(folder_path, "PNG")
    if not os.path.isdir(png_dir):
        missing.append((folder, "无 PNG 目录"))
        continue
    # 匹配 *-Harm_0.png（不区分大小写）
    hits = glob.glob(os.path.join(png_dir, "*-Harm_0.png")) or \
           glob.glob(os.path.join(png_dir, "*-harm_0.png"))
    if hits:
        rows.append((folder, hits[0]))
    else:
        missing.append((folder, "未找到 Harm_0"))

print(f"共 {len(folders)} 个文件夹，其中 {len(rows)} 个找到 Harm_0 图片")
if missing:
    print("未找到 Harm_0 的文件夹：")
    for name, reason in missing:
        print(f"  - {name}: {reason}")

# 创建 XLSX
wb = Workbook()
ws = wb.active
ws.title = "角色形象对照"

# 表头
ws["A1"] = "角色名"
ws["B1"] = "Harm_0 形象"
ws.column_dimensions["A"].width = 20
ws.column_dimensions["B"].width = 30

# 估算图片高度，预留空间
cell_height_px = 140          # 每个格子预留的图片区域高度（像素）
target_width = 220            # 目标显示宽度（像素）

row_idx = 2
for name, img_path in rows:
    ws.cell(row=row_idx, column=1, value=name)
    try:
        img = XLImage(img_path)
        # 按目标宽度等比缩放，防止图片过大
        scale = target_width / img.width if img.width > target_width else 1.0
        img.width = int(img.width * scale)
        img.height = int(img.height * scale)
        img.anchor = f"B{row_idx}"
        ws.add_image(img)
        # 设置行高，确保图片可完整显示
        ws.row_dimensions[row_idx].height = max(img.height + 10, 60)
    except Exception as e:
        print(f"  加载图片失败 {name}: {e}")
        ws.row_dimensions[row_idx].height = 60
    row_idx += 1

wb.save(OUTPUT)
print(f"已生成：{OUTPUT}")
print(f"共写入 {len(rows)} 个角色")
